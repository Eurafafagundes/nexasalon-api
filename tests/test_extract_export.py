"""Testes HTTP de `GET /api/v1/extract/export` (Etapa L, Bloco 12) —
exportação `.xlsx` REAL (nunca CSV disfarçado) do Financeiro > Extrato.
Cobre: o arquivo devolvido é um `.xlsx` de verdade (abre com `openpyxl`)
com as MESMAS vendas que `GET /extract` mostraria; o filtro de período é
respeitado; a mesma permissão (`finance.view`) e o mesmo isolamento
multi-tenant do `GET /extract` original — nunca uma segunda rota com
regra de acesso diferente pro mesmo dado."""
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

_START = "2026-08-13T14:00:00-03:00"  # quinta-feira


def _setup_closed_order(c):
    """Cria unidade/profissional/serviço/cliente/agendamento/comanda via
    API e fecha com um pagamento — mesmo padrão de
    `test_orders_api.py::_setup_finished_appointment`."""
    branch = c.post("/api/v1/branches", json={"name": "Matriz", "slug": f"matriz-{uuid.uuid4().hex[:6]}"}).json()
    professional = c.post("/api/v1/professionals", json={"name": "Ianka"}).json()
    service = c.post(
        "/api/v1/services", json={"name": "Corte", "default_duration_minutes": 60, "default_price": "150.00"}
    ).json()
    c.put(f"/api/v1/professionals/{professional['id']}/services", json={"items": [{"service_id": service["id"]}]})
    c.put(
        f"/api/v1/professionals/{professional['id']}/working-hours",
        json={"items": [{"weekday": 4, "start_time": "09:00:00", "end_time": "20:00:00"}]},
    )
    client = c.post("/api/v1/clients", json={"name": "Cliente Exportação"}).json()
    appt = c.post(
        "/api/v1/appointments",
        json={
            "branch_id": branch["id"], "client_id": client["id"],
            "items": [{"professional_id": professional["id"], "service_id": service["id"], "start_at": _START}],
        },
    ).json()
    for target in ["confirmed", "waiting", "in_progress", "finished"]:
        assert c.patch(f"/api/v1/appointments/{appt['id']}/status", json={"status": target}).status_code == 200

    register = c.post("/api/v1/cash-registers", json={"branch_id": branch["id"], "initial_amount": "0"}).json()
    order = c.post("/api/v1/orders", json={"appointment_id": appt["id"]}).json()
    closed = c.post(
        f"/api/v1/orders/{order['id']}/close",
        json={"payments": [{"method": "pix", "amount": "150.00", "cash_register_id": register["id"]}]},
    )
    assert closed.status_code == 200, closed.text
    return client


def test_export_gera_xlsx_real_com_a_mesma_venda_do_extrato(client_as, org_a_actor):
    c = client_as(org_a_actor)
    client = _setup_closed_order(c)

    resp = c.get("/api/v1/extract/export")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith('.xlsx"')

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header, *data_rows = rows
    assert header[0] == "Data"
    assert "Valor" in header

    client_col = header.index("Cliente")
    value_col = header.index("Valor")
    matching = [r for r in data_rows if r[client_col] == client["name"]]
    assert len(matching) == 1
    assert matching[0][value_col] == 150.0


def test_export_respeita_filtro_de_periodo(client_as, org_a_actor):
    c = client_as(org_a_actor)
    _setup_closed_order(c)

    far_future = (datetime.now(timezone.utc) + timedelta(days=365)).isoformat()
    resp = c.get("/api/v1/extract/export", params={"date_from": far_future})
    assert resp.status_code == 200, resp.text
    workbook = load_workbook(BytesIO(resp.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert len(rows) == 1  # só o header — a venda de hoje fica fora da janela futura

    resp_all = c.get("/api/v1/extract/export")
    rows_all = list(load_workbook(BytesIO(resp_all.content)).active.iter_rows(values_only=True))
    assert len(rows_all) == 2  # header + a venda


def test_export_exige_finance_view(client_as, org_a_actor):
    import dataclasses

    restricted = dataclasses.replace(org_a_actor, permissions=org_a_actor.permissions - {"finance.view"})
    resp = client_as(restricted).get("/api/v1/extract/export")
    assert resp.status_code == 403


def test_export_nao_vaza_entre_organizacoes(client_as, org_a_actor, org_b_actor):
    c_a = client_as(org_a_actor)
    client_a = _setup_closed_order(c_a)

    c_b = client_as(org_b_actor)
    resp = c_b.get("/api/v1/extract/export")
    assert resp.status_code == 200, resp.text
    workbook = load_workbook(BytesIO(resp.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    header, *data_rows = rows
    client_col = header.index("Cliente")
    assert all(r[client_col] != client_a["name"] for r in data_rows)
