"""Extrato — Financeiro > Extrato (item 17/18/19 da rodada "evolução
funcional"). Unidade principal é a COMANDA (`Order` fechada), nunca o
item/pagamento: uma comanda com 2 serviços ou 2 pagamentos continua
sendo UMA linha (item "pagamento misto não duplica faturamento").
Movimentações (`CashMovement`) são uma granularidade SEPARADA — nunca
somadas junto com o faturamento de vendas na mesma conta (item
"Movimentações").
"""
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from sqlalchemy.orm import Session

from nexasalon_api.core.actor import ActorContext
from nexasalon_api.models.cash_register import CashMovement
from nexasalon_api.models.enums import CashMovementType, OrderStatus
from nexasalon_api.models.order import Order
from nexasalon_api.repositories import (
    branch_repo,
    cash_movement_repo,
    cash_register_repo,
    client_repo,
    order_repo,
)


@dataclass
class ExtractSummary:
    revenue_total: Decimal = Decimal(0)  # Receitas — soma de comandas FECHADAS no período.
    expense_total: Decimal = Decimal(0)  # Despesas — soma de CashMovement WITHDRAWAL no período.
    result: Decimal = Decimal(0)
    sales: list[Order] = field(default_factory=list)
    movements: list[CashMovement] = field(default_factory=list)
    client_names: dict[uuid.UUID, str] = field(default_factory=dict)


def get_extract(
    session: Session,
    actor: ActorContext,
    *,
    date_from: datetime | None,
    date_to: datetime | None,
    status: OrderStatus | None = None,
) -> ExtractSummary:
    sales = order_repo.list_for_org(
        session, actor.organization_id, status=status, date_from=date_from, date_to=date_to
    )
    movements = cash_movement_repo.list_for_org(
        session, actor.organization_id, date_from=date_from, date_to=date_to
    )

    revenue_total = sum(
        (sum((item.price for item in o.items), Decimal(0)) for o in sales if o.status == OrderStatus.CLOSED),
        Decimal(0),
    )
    expense_total = sum(
        (m.amount for m in movements if m.type == CashMovementType.WITHDRAWAL), Decimal(0)
    )

    client_ids = {o.client_id for o in sales}
    client_names = {
        c.id: c.name for c in (client_repo.get(session, actor.organization_id, cid) for cid in client_ids) if c
    }

    return ExtractSummary(
        revenue_total=revenue_total,
        expense_total=expense_total,
        result=revenue_total - expense_total,
        sales=sales,
        movements=movements,
        client_names=client_names,
    )


# ---------------------------------------------------------------------
# Etapa L, Bloco 12 — exportação .xlsx. Reaproveita `get_extract`
# INTEGRALMENTE (mesmo filtro, mesmo RBAC — `finance.view` continua
# sendo o único portão, checado na rota) — o conteúdo do arquivo é
# EXATAMENTE o mesmo dado que o Extrato já mostra na tela, nunca uma
# segunda consulta paralela que poderia divergir. Granularidade
# permanece "uma linha = uma Comanda" pra vendas (mesma decisão de
# `ExtractSaleRow`, ver docstring do módulo) + "uma linha = uma
# movimentação manual" — nunca uma linha por item/pagamento (evita
# duplicar faturamento na planilha, item 18 do pedido original).
# ---------------------------------------------------------------------

_HEADER = [
    "Data",
    "Hora",
    "Tipo",
    "Cliente",
    "Comanda",
    "Descrição",
    "Serviço/Produto",
    "Profissional",
    "Forma de pagamento",
    "Valor",
    "Caixa",
    "Unidade",
    "Status",
]


def _sale_row(order: Order, client_name: str, branch_name: str) -> list:
    total = sum((item.price for item in order.items), Decimal(0)) + sum(
        (item.unit_price * item.quantity for item in order.product_items), Decimal(0)
    )
    services_summary = " + ".join(dict.fromkeys(i.service_name for i in order.items)) or "—"
    products_summary = " + ".join(dict.fromkeys(i.product_name for i in order.product_items))
    item_summary = " + ".join(part for part in (services_summary, products_summary) if part) or "—"
    professionals_summary = " + ".join(dict.fromkeys(i.professional_name for i in order.items)) or "—"
    payment_summary = " + ".join(dict.fromkeys(p.method.value for p in order.payments)) or "—"
    moment = order.closed_at or order.created_at
    return [
        moment.date().isoformat(),
        moment.strftime("%H:%M"),
        "Venda",
        client_name,
        f"#{order.order_number}",
        item_summary,
        item_summary,
        professionals_summary,
        payment_summary,
        float(total),
        "—",
        branch_name,
        order.status.value,
    ]


def _movement_row(movement: CashMovement, branch_name: str) -> list:
    kind = "Entrada" if movement.type == CashMovementType.SUPPLY else "Despesa"
    signed_amount = movement.amount if movement.type == CashMovementType.SUPPLY else -movement.amount
    return [
        movement.created_at.date().isoformat(),
        movement.created_at.strftime("%H:%M"),
        kind,
        "—",
        "—",
        movement.description,
        movement.category or "—",
        "—",
        movement.method.value,
        float(signed_amount),
        "Caixa",
        branch_name,
        "—",
    ]


def build_extract_workbook(
    session: Session,
    actor: ActorContext,
    *,
    date_from: datetime | None,
    date_to: datetime | None,
    status: OrderStatus | None,
) -> bytes:
    """Gera o `.xlsx` no BACKEND (item explícito do pedido: "preferir
    geração no backend se isso garantir RBAC, consistência e permitir
    volumes maiores") — nunca um CSV disfarçado de xlsx (usa `openpyxl`,
    formato real do Excel)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    summary = get_extract(session, actor, date_from=date_from, date_to=date_to, status=status)

    branch_cache: dict[uuid.UUID, str] = {}

    def branch_name(branch_id: uuid.UUID | None) -> str:
        if branch_id is None:
            return "—"
        if branch_id not in branch_cache:
            branch = branch_repo.get(session, actor.organization_id, branch_id)
            branch_cache[branch_id] = branch.name if branch is not None else "—"
        return branch_cache[branch_id]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extrato"
    sheet.append(_HEADER)

    for order in summary.sales:
        client_name = summary.client_names.get(order.client_id, "Cliente removido")
        sheet.append(_sale_row(order, client_name, branch_name(order.branch_id)))

    for movement in summary.movements:
        register = cash_register_repo.get(session, actor.organization_id, movement.cash_register_id)
        sheet.append(_movement_row(movement, branch_name(register.branch_id if register else None)))

    for idx, header in enumerate(_HEADER, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(len(header) + 2, 14)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_extract_filename(reference: datetime) -> str:
    return f"extrato-nexasalon-{reference.strftime('%Y-%m')}.xlsx"
